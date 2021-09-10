import csv
from openpyxl import Workbook
from openpyxl import load_workbook

def output_by_subject(reader):
    sub_dic={}
    for line in reader:
        roll_number=line['rollno']
        sub_number=line['subno']
        if(sub_number not in sub_dic):
            wb=Workbook()
            sheet=wb.active
            sheet['A1']="rollno"
            sheet['B1']='register_sem'
            sheet['C1']='subno'
            sheet['D1']='sub_type'
            wb.save('output_by_subject\\'+ sub_number +'.xlsx')
            sub_dic[sub_number]=2
        row=str(sub_dic[sub_number])  
        w=load_workbook('output_by_subject\\'+sub_number+'.xlsx')
        Sheet=w.active
        Sheet['A'+row]=roll_number
        Sheet['B'+row]=line['register_sem']
        Sheet['C'+row]=line['subno']
        Sheet['D'+row]=line['sub_type']
        w.save('output_by_subject\\'+sub_number+'.xlsx')
        sub_dic[sub_number]+=1

    return

def output_individual_roll(reader):
    roll_dic={}
    for line in reader:
            roll_number=line['rollno']
            if(roll_number not in roll_dic):
                wb=Workbook()
                sheet=wb.active
                sheet['A1']="rollno"
                sheet['B1']='register_sem'
                sheet['C1']='subno'
                sheet['D1']='sub_type'
                wb.save('output_individual_roll\\'+ roll_number +'.xlsx')
                roll_dic[roll_number]=2
            row=str(roll_dic[roll_number])  
            w=load_workbook('output_individual_roll\\'+roll_number+'.xlsx')
            Sheet=w.active
            Sheet['A'+row]=roll_number
            Sheet['B'+row]=line['register_sem']
            Sheet['C'+row]=line['subno']
            Sheet['D'+row]=line['sub_type']
            w.save('output_individual_roll\\'+roll_number+'.xlsx')
            roll_dic[roll_number]+=1
           

    return

with open("regtable_old.csv","r") as file:
    reader=csv.DictReader(file)
    output_individual_roll(reader)
    output_by_subject(reader)
    
