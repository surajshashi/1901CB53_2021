from openpyxl import Workbook
from openpyxl import load_workbook
import csv
import os

#Dictionary converting grades into numericals
gradeDict={'AA':10, 'AB':9, 'BB':8, 'BC':7, 'CC':6, 'CD':5, 'DD':4, 'F':0,'F*':0, 'DD*':4, 'I':0} 

#Arranging all grade details of Students to a dictionary
def gradeDetails():
    with open('grades.csv','r') as file:
        reader=csv.DictReader(file)
        gradDet={}
        for row in reader:
            rollnum=str(row['Roll'])
            sem=int(row['Sem'])
            grade=gradeDict[row['Grade'].strip()]
            crd=int(row['Credit'])
            score=int(crd*grade)
            if(rollnum not in gradDet):
                gradDet[rollnum]=[[0,0],[0,0],[0,0],[0,0],[0,0],[0,0],[0,0],[0,0],[0,0],[0,0],[0,0]]
            
            gradDet[rollnum][sem][0]=gradDet[rollnum][sem][0]+crd
            
            gradDet[rollnum][sem][1]=gradDet[rollnum][sem][1]+ score
            
    return gradDet

#Making a Dictionery of Subject Details
def subjectDetails():
    with open('subjects_master.csv','r') as file:
        reader=csv.DictReader(file)
        subjDict={}
        for row in reader:
            subno=row['subno']
            subname=row['subname']
            ltp=row['ltp']
            crd=row['crd']
            subjDict[subno]=[subname,ltp,crd]
    return subjDict

#Creating marksheet of students based on their roll number
def generate_marksheet(subjDict,gradDet):

    #Creating the file and overall Score sheet 
    with open("names-roll.csv","r") as file:
        reader=csv.DictReader(file)
        
        for row in reader:
            roll_number=str(row['Roll'])
            name=row['Name']
            discipline=roll_number[4:6]
            wb=Workbook()
            sheet=wb.active
            sheet['A1']='Roll No.'
            sheet['A2']='Name of Student'
            sheet['A3']='Discipline'
            sheet['A4']='Semester No.'
            sheet['A5']='Semester wise Credit Taken'
            sheet['A6']='SPI'
            sheet['A7']='Total Credits Taken'
            sheet['A8']='CPI'
            sheet['B1']=roll_number
            sheet['B2']=name
            sheet['B3']=discipline
            sheet.title='Overall'

    #Crating file
            wb.save('output\\' + roll_number +'.xlsx')
            
    
    
    semCol={'1':'B','2':'C','3':'D', '4':'E','5':'F','6':'G','7':'H','8':'I','9':'J','10':'K'}

    #Creating the Semester Wise sheets of each students
    with open('grades.csv','r') as file:
        reader=csv.DictReader(file)
      
        for row in reader:
            roll_number=str(row['Roll'])
            sem=row['Sem']
            subCode=row['SubCode']
            grade=row['Grade']
            subType=row['Sub_Type']
            credit=int(row['Credit'])
            wb=load_workbook('output\\'+roll_number+'.xlsx')
            try:
                sheet=wb['Sem'+sem]
                row=str(sheet.max_row + 1)
                sheet['A'+row]=sheet.max_row
                sheet['B'+row]=subCode
                sheet['C'+row]=subjDict[subCode][0]
                sheet['D'+row]=subjDict[subCode][1]
                sheet['E'+row]=credit
                sheet['F'+row]=subType
                sheet['G'+row]=grade

            except:
                
                wb.create_sheet(index=int(sem),title='Sem'+str(sem))
                sheet=wb['Sem'+ sem]
                heading=('Sl No.',	'Subject No.',	'Subject Name', 'L-T-P','Credit '	,'Subject Type',	'Grade')
                sheet.append(heading)
                wb.save('output\\'+roll_number+'.xlsx')
                wb=load_workbook('output\\'+roll_number+'.xlsx')
                sheet=wb['Sem'+sem]
                row=str(sheet.max_row + 1)
                sheet['A'+row]=sheet.max_row
                sheet['B'+row]=subCode
                sheet['C'+row]=subjDict[subCode][0]
                sheet['D'+row]=subjDict[subCode][1]
                sheet['E'+row]=credit
                sheet['F'+row]=subType
                sheet['G'+row]=grade


            sheet=wb['Overall']
            if(sem=='10'):
                sheet[semCol['9']+'4']=10
            else:
                sheet[semCol[sem]+'4']=int(sem)

    #Inserting the Semester wise credit taken
            if(sem=='10'):
                sheet[semCol['9']+'5']=gradDet[roll_number][10][0]
            else:
                sheet[semCol[sem]+'5']=gradDet[roll_number][int(sem)][0]

    #Inserting the SPI
            if(sem=='10'):
                sheet[semCol['9']+'6']=round(gradDet[roll_number][10][1]/gradDet[roll_number][10][0],2)
            else:
                sheet[semCol[sem]+'6']=round(gradDet[roll_number][int(sem)][1]/gradDet[roll_number][int(sem)][0],2)

    #Inserting the CPI and total credits taken
            if(sem=='1'):
                sheet[semCol[sem]+'7']= gradDet[roll_number][int(sem)][0]
                sheet[semCol[sem]+'8']=round(gradDet[roll_number][int(sem)][1]/gradDet[roll_number][int(sem)][0],2)
            elif(sem=='10'):
                sheet[semCol['9']+'7']= gradDet[roll_number][10][0]
                sheet[semCol['9']+'8']= round((gradDet[roll_number][10][1] + float(sheet[semCol['8']+'8'].value)*int(sheet[semCol['8']+'7'].value))/(gradDet[roll_number][10][0] + int(sheet[semCol['8']+'7'].value) ),2)
            else:
                crdt=gradDet[roll_number][int(sem)][0]
                spi=round(gradDet[roll_number][int(sem)][1]/gradDet[roll_number][int(sem)][0],2)
    #Total Credits
                sheet[semCol[sem]+'7']=int(sheet[semCol[str(int(sem)-1)]+'7'].value) + gradDet[roll_number][int(sem)][0]
                totalcred=int(sheet[semCol[str(int(sem)-1)]+'7'].value) + gradDet[roll_number][int(sem)][0]
    #CPI
                sheet[semCol[sem]+'8']=round((float(sheet[semCol[str(int(sem)-1)]+'8'].value)*float(sheet[semCol[str(int(sem)-1)]+'7'].value) + crdt*spi)/totalcred,2)

            wb.save('output\\'+roll_number+'.xlsx')
        



    return

#Creating the Output Directory
try:
    os.mkdir('output')
except:
    pass

#Calling the Subject Details Function which return dictionary
sub=subjectDetails()

#Calling the Grade Details Function which return dictionary
grd=gradeDetails()

#Calling Funtion to generate Marksheet
generate_marksheet(sub,grd)



